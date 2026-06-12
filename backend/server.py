from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import uuid
import secrets
import string
import logging
import asyncio
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

import bcrypt
import jwt
import resend
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, EmailStr
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ---------- Setup ----------
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger("trip-splitter")

mongo_url = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")
APP_URL = os.environ.get("APP_URL", "")

if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

client = AsyncIOMotorClient(mongo_url)
db = client[DB_NAME]

app = FastAPI(title="Trip Splitter")
api = APIRouter(prefix="/api")

CATEGORIES = ["Travel", "Accommodation", "Local Transportation",
              "Local Sightseeing", "Food", "Shopping", "Other"]


# ---------- Utils ----------
def hash_secret(v: str) -> str:
    return bcrypt.hashpw(v.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_secret(v: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(v.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id, "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(days=30),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")


async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    payload = decode_token(authorization[7:])
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0, "pin_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user


def gen_trip_code() -> str:
    return ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))


def gen_id() -> str:
    return str(uuid.uuid4())


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def iso(dt) -> Optional[str]:
    if dt is None:
        return None
    if isinstance(dt, str):
        return dt
    return dt.isoformat()


# ---------- Models ----------
class RegisterIn(BaseModel):
    email: EmailStr
    pin: str = Field(min_length=4, max_length=4)
    name: str = Field(min_length=1)
    password: Optional[str] = None  # legacy, optional


class LoginIn(BaseModel):
    email: EmailStr
    password: Optional[str] = None
    pin: Optional[str] = None


class ForgotIn(BaseModel):
    email: EmailStr


class ResetPinIn(BaseModel):
    token: str
    new_pin: str = Field(min_length=4, max_length=4)


class TripIn(BaseModel):
    name: str
    travel_date: str  # DD-MM-YY
    budget: Optional[float] = None
    currency: str = "INR"


class TripUpdate(BaseModel):
    name: Optional[str] = None
    travel_date: Optional[str] = None
    budget: Optional[float] = None
    currency: Optional[str] = None


class MemberIn(BaseModel):
    name: str
    kind: Literal["individual", "family"] = "individual"
    family_members: List[str] = []  # names of family members
    email: Optional[EmailStr] = None  # optional email to auto-link an app user


class MemberUpdate(BaseModel):
    name: Optional[str] = None
    kind: Optional[Literal["individual", "family"]] = None
    family_members: Optional[List[str]] = None
    email: Optional[str] = None  # can be empty string to clear
    reweight_past: Optional[bool] = True  # if False, snapshot old weights onto past expenses


class ExpenseIn(BaseModel):
    kind: Literal["expense", "income"] = "expense"
    amount: float
    category: str
    description: Optional[str] = ""
    date: str  # DD-MM-YY
    paid_by_member_id: str  # member id (individual or family) who paid
    split_member_ids: List[str] = []  # if empty, split among all
    weight_snapshots: Optional[dict] = None  # member_id -> custom weight (e.g. partial family)
    receipt_base64: Optional[str] = None


class ExpenseUpdate(BaseModel):
    kind: Optional[Literal["expense", "income"]] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    paid_by_member_id: Optional[str] = None
    split_member_ids: Optional[List[str]] = None
    weight_snapshots: Optional[dict] = None
    receipt_base64: Optional[str] = None
    force: Optional[bool] = False


class SettleIn(BaseModel):
    from_member_id: str
    to_member_id: str
    amount: float


# ---------- Auth ----------
@api.post("/auth/register")
async def register(body: RegisterIn):
    email = body.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already registered")
    if not body.pin.isdigit():
        raise HTTPException(400, "PIN must be 4 digits")
    uid = gen_id()
    # Password is optional now; PIN is the primary credential.
    password_hash = hash_secret(body.password) if body.password else hash_secret(secrets.token_urlsafe(16))
    doc = {
        "id": uid, "email": email, "name": body.name,
        "password_hash": password_hash,
        "pin_hash": hash_secret(body.pin),
        "role": "user",
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(doc)
    token = create_token(uid, email)
    return {"access_token": token, "user": {"id": uid, "email": email, "name": body.name, "role": "user"}}


@api.post("/auth/login")
async def login(body: LoginIn):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(401, "Invalid credentials")
    if body.password:
        if not verify_secret(body.password, user["password_hash"]):
            raise HTTPException(401, "Invalid credentials")
    elif body.pin:
        if not verify_secret(body.pin, user["pin_hash"]):
            raise HTTPException(401, "Invalid PIN")
    else:
        raise HTTPException(400, "Provide password or pin")
    token = create_token(user["id"], email)
    return {"access_token": token,
            "user": {"id": user["id"], "email": email, "name": user["name"], "role": user.get("role", "user")}}


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


@api.post("/auth/forgot-pin")
async def forgot_pin(body: ForgotIn):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if user:
        token = secrets.token_urlsafe(32)
        await db.password_reset_tokens.insert_one({
            "token": token, "user_id": user["id"], "kind": "pin",
            "expires_at": (now_utc() + timedelta(hours=1)).isoformat(),
            "used": False,
        })
        reset_link = f"{APP_URL}/reset-pin?token={token}" if APP_URL else token
        logger.info(f"[PIN RESET] Email={email}  Token={token}  Link={reset_link}")
        if RESEND_API_KEY:
            try:
                await asyncio.to_thread(resend.Emails.send, {
                    "from": SENDER_EMAIL, "to": [email],
                    "subject": "Reset your Trip Splitter PIN",
                    "html": (
                        f"<div style='font-family:sans-serif'>"
                        f"<h2>Reset your PIN</h2>"
                        f"<p>Hi {user.get('name','there')},</p>"
                        f"<p>Tap the link below to reset your 4-digit PIN. This link expires in 1 hour.</p>"
                        f"<p><a href='{reset_link}' style='background:#1C3F39;color:#fff;padding:12px 20px;border-radius:24px;text-decoration:none;display:inline-block'>Reset PIN</a></p>"
                        f"<p>Or use this token in the app: <b>{token}</b></p>"
                        f"<p style='color:#888;font-size:12px'>If you didn't request this, ignore this email.</p>"
                        f"</div>"
                    ),
                })
            except Exception as e:
                logger.warning(f"Resend send failed: {e}")
    return {"ok": True, "message": "If this email exists, a reset link has been sent."}


@api.post("/auth/reset-pin")
async def reset_pin(body: ResetPinIn):
    if not body.new_pin.isdigit():
        raise HTTPException(400, "PIN must be 4 digits")
    rec = await db.password_reset_tokens.find_one({"token": body.token})
    if not rec or rec.get("used"):
        raise HTTPException(400, "Invalid or already-used token")
    if datetime.fromisoformat(rec["expires_at"]) < now_utc():
        raise HTTPException(400, "Token expired")
    await db.users.update_one({"id": rec["user_id"]},
                              {"$set": {"pin_hash": hash_secret(body.new_pin)}})
    await db.password_reset_tokens.update_one({"token": body.token}, {"$set": {"used": True}})
    return {"ok": True}


# Backward-compat aliases (kept so old frontend builds keep working)
@api.post("/auth/forgot-password")
async def forgot_password_alias(body: ForgotIn):
    return await forgot_pin(body)


# ---------- Trips ----------
async def _trip_or_404(trip_id: str, user_id: str) -> dict:
    trip = await db.trips.find_one({"id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(404, "Trip not found")
    if user_id not in trip.get("user_ids", []):
        raise HTTPException(403, "Not a member of this trip")
    return trip


@api.post("/trips")
async def create_trip(body: TripIn, user=Depends(get_current_user)):
    tid = gen_id()
    code = gen_trip_code()
    while await db.trips.find_one({"code": code}):
        code = gen_trip_code()
    # create an "owner member" automatically (individual)
    owner_member = {
        "id": gen_id(), "name": user["name"], "kind": "individual",
        "family_members": [], "email": user["email"], "user_id": user["id"],
    }
    doc = {
        "id": tid, "code": code, "name": body.name, "travel_date": body.travel_date,
        "budget": body.budget, "currency": body.currency or "INR",
        "owner_id": user["id"], "user_ids": [user["id"]],
        "members": [owner_member],
        "created_at": now_utc().isoformat(),
    }
    await db.trips.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/trips")
async def list_trips(user=Depends(get_current_user)):
    cur = db.trips.find({"user_ids": user["id"]}, {"_id": 0}).sort("created_at", -1)
    return await cur.to_list(500)


@api.get("/trips/{trip_id}")
async def get_trip(trip_id: str, user=Depends(get_current_user)):
    return await _trip_or_404(trip_id, user["id"])


@api.patch("/trips/{trip_id}")
async def update_trip(trip_id: str, body: TripUpdate, user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        await db.trips.update_one({"id": trip_id}, {"$set": updates})
    return await db.trips.find_one({"id": trip_id}, {"_id": 0})


@api.delete("/trips/{trip_id}")
async def delete_trip(trip_id: str, user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    if trip["owner_id"] != user["id"]:
        raise HTTPException(403, "Only the owner can delete")
    await db.trips.delete_one({"id": trip_id})
    await db.expenses.delete_many({"trip_id": trip_id})
    await db.settlements.delete_many({"trip_id": trip_id})
    return {"ok": True}


@api.post("/trips/join")
async def join_trip(body: dict, user=Depends(get_current_user)):
    code = (body.get("code") or "").upper().strip()
    trip = await db.trips.find_one({"code": code}, {"_id": 0})
    if not trip:
        raise HTTPException(404, "Trip not found")
    if user["id"] in trip.get("user_ids", []):
        return trip
    # Check if a family member in this trip has this user's email -> link instead of adding
    user_email = user["email"].lower().strip()
    linked_family = None
    for m in trip.get("members", []):
        if m.get("kind") == "family" and (m.get("email") or "").lower() == user_email and not m.get("user_id"):
            linked_family = m
            break
    if linked_family:
        await db.trips.update_one(
            {"id": trip["id"], "members.id": linked_family["id"]},
            {"$push": {"user_ids": user["id"]},
             "$set": {"members.$.user_id": user["id"]}},
        )
    else:
        new_member = {
            "id": gen_id(), "name": user["name"], "kind": "individual",
            "family_members": [], "email": user_email, "user_id": user["id"],
        }
        await db.trips.update_one(
            {"id": trip["id"]},
            {"$push": {"user_ids": user["id"], "members": new_member}},
        )
    return await db.trips.find_one({"id": trip["id"]}, {"_id": 0})


# ---------- Members ----------
@api.post("/trips/{trip_id}/members")
async def add_member(trip_id: str, body: MemberIn, user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Name is required")
    # Duplicate-name check (case-insensitive) per trip
    for m in trip.get("members", []):
        if m["name"].lower() == name.lower():
            raise HTTPException(400, f"A member named '{name}' already exists in this trip")
    email = (body.email or "").lower().strip() or None
    # Determine if this email matches an existing user-individual we should merge with (in-place)
    merge_target = None
    if email and body.kind == "family":
        for m in trip.get("members", []):
            if (m.get("email") or "").lower() == email and m.get("user_id") and m.get("kind") == "individual":
                merge_target = m; break
    # Duplicate email check (skip if we plan to merge that one)
    if email:
        for m in trip.get("members", []):
            if merge_target and m["id"] == merge_target["id"]:
                continue
            if (m.get("email") or "").lower() == email:
                raise HTTPException(400, f"A member with email '{email}' already exists in this trip")
    new_member = {
        "id": gen_id(), "name": name, "kind": body.kind,
        "family_members": body.family_members if body.kind == "family" else [],
        "email": email, "user_id": None,
    }
    # If a user already in the trip has this email AND currently exists as an individual,
    # convert that individual into this family IN-PLACE (preserves member.id so all past
    # expenses automatically apply to the family — avoids double-counting).
    if merge_target:
        await db.trips.update_one(
            {"id": trip_id, "members.id": merge_target["id"]},
            {"$set": {
                "members.$.name": name,
                "members.$.kind": "family",
                "members.$.family_members": body.family_members,
                "members.$.email": email,
            }},
        )
        t = await db.trips.find_one({"id": trip_id}, {"_id": 0})
        return next((m for m in t["members"] if m["id"] == merge_target["id"]), None)
    await db.trips.update_one({"id": trip_id}, {"$push": {"members": new_member}})
    return new_member


@api.patch("/trips/{trip_id}/members/{member_id}")
async def update_member(trip_id: str, member_id: str, body: MemberUpdate, user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    target = next((m for m in trip["members"] if m["id"] == member_id), None)
    if not target:
        raise HTTPException(404, "Member not found")
    updates: dict = {}
    if body.name is not None:
        nm = body.name.strip()
        if not nm:
            raise HTTPException(400, "Name cannot be empty")
        # duplicate check excluding self
        for m in trip["members"]:
            if m["id"] != member_id and m["name"].lower() == nm.lower():
                raise HTTPException(400, f"A member named '{nm}' already exists in this trip")
        updates["members.$.name"] = nm
    if body.kind is not None:
        updates["members.$.kind"] = body.kind
    new_kind = body.kind if body.kind is not None else target["kind"]
    new_fm = body.family_members if body.family_members is not None else target.get("family_members", [])
    if new_kind != "family":
        new_fm = []
    if body.family_members is not None or body.kind is not None:
        updates["members.$.family_members"] = new_fm
    if body.email is not None:
        em = (body.email or "").lower().strip() or None
        if em:
            for m in trip["members"]:
                if m["id"] != member_id and (m.get("email") or "").lower() == em:
                    raise HTTPException(400, f"A member with email '{em}' already exists in this trip")
        updates["members.$.email"] = em

    # If family members list changed and user chose NOT to re-weight past, snapshot old weights
    old_fm = target.get("family_members", [])
    old_weight = _weight_of_member(target)
    new_weight_member = {**target, "kind": new_kind, "family_members": new_fm}
    new_weight = _weight_of_member(new_weight_member)
    if old_weight != new_weight and body.reweight_past is False:
        # For every past expense that has this member in split_member_ids, snapshot the OLD weight
        async for e in db.expenses.find({"trip_id": trip_id, "split_member_ids": member_id}):
            snap = e.get("weight_snapshots") or {}
            if member_id not in snap:
                snap[member_id] = old_weight
                await db.expenses.update_one({"id": e["id"]}, {"$set": {"weight_snapshots": snap}})

    if updates:
        await db.trips.update_one({"id": trip_id, "members.id": member_id}, {"$set": updates})
    t = await db.trips.find_one({"id": trip_id}, {"_id": 0})
    return next((m for m in t["members"] if m["id"] == member_id), None)


@api.delete("/trips/{trip_id}/members/{member_id}")
async def delete_member(trip_id: str, member_id: str, user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    # cannot remove if member appears in any expense
    exists = await db.expenses.find_one({"trip_id": trip_id,
                                         "$or": [{"paid_by_member_id": member_id},
                                                 {"split_member_ids": member_id}]})
    if exists:
        raise HTTPException(400, "Member has expenses; cannot delete")
    await db.trips.update_one({"id": trip_id}, {"$pull": {"members": {"id": member_id}}})
    return {"ok": True}


# ---------- Expenses ----------
def _weight_of_member(m: dict) -> int:
    if m["kind"] == "family":
        return max(1, len(m.get("family_members", [])))
    return 1


@api.post("/trips/{trip_id}/expenses")
async def add_expense(trip_id: str, body: ExpenseIn, force: bool = False,
                      user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    if body.category not in CATEGORIES:
        raise HTTPException(400, "Invalid category")
    member_ids = {m["id"] for m in trip["members"]}
    if body.paid_by_member_id not in member_ids:
        raise HTTPException(400, "paid_by_member_id invalid")
    split_ids = body.split_member_ids or [m["id"] for m in trip["members"]]
    for sid in split_ids:
        if sid not in member_ids:
            raise HTTPException(400, f"split member {sid} invalid")

    # budget over-check (category vs overall)
    warning = None
    if body.kind == "expense" and trip.get("budget"):
        cur = await db.expenses.aggregate([
            {"$match": {"trip_id": trip_id, "kind": "expense"}},
            {"$group": {"_id": None, "sum": {"$sum": "$amount"}}},
        ]).to_list(1)
        current = cur[0]["sum"] if cur else 0
        if current + body.amount > trip["budget"]:
            warning = f"This expense puts you {(current + body.amount) - trip['budget']:.2f} {trip.get('currency','INR')} over the trip budget."
            if not force:
                return {"requires_confirmation": True, "warning": warning}

    eid = gen_id()
    doc = {
        "id": eid, "trip_id": trip_id, "kind": body.kind,
        "amount": float(body.amount), "category": body.category,
        "description": body.description or "",
        "date": body.date, "paid_by_member_id": body.paid_by_member_id,
        "split_member_ids": split_ids,
        "weight_snapshots": body.weight_snapshots or None,
        "receipt_base64": body.receipt_base64,
        "created_by": user["id"], "created_at": now_utc().isoformat(),
    }
    await db.expenses.insert_one(doc)
    doc.pop("_id", None)
    return {"expense": doc, "warning": warning}


@api.get("/trips/{trip_id}/expenses")
async def list_expenses(trip_id: str, user=Depends(get_current_user)):
    await _trip_or_404(trip_id, user["id"])
    cur = db.expenses.find({"trip_id": trip_id}, {"_id": 0}).sort("created_at", -1)
    return await cur.to_list(1000)


@api.patch("/trips/{trip_id}/expenses/{expense_id}")
async def update_expense(trip_id: str, expense_id: str, body: ExpenseUpdate,
                         user=Depends(get_current_user)):
    await _trip_or_404(trip_id, user["id"])
    updates = {k: v for k, v in body.model_dump().items() if v is not None and k != "force"}
    if not updates:
        return await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    await db.expenses.update_one({"id": expense_id, "trip_id": trip_id}, {"$set": updates})
    return await db.expenses.find_one({"id": expense_id}, {"_id": 0})


@api.delete("/trips/{trip_id}/expenses/{expense_id}")
async def delete_expense(trip_id: str, expense_id: str, user=Depends(get_current_user)):
    await _trip_or_404(trip_id, user["id"])
    await db.expenses.delete_one({"id": expense_id, "trip_id": trip_id})
    return {"ok": True}


# ---------- Balances / Settle Up ----------
async def _compute_balances(trip_id: str) -> dict:
    trip = await db.trips.find_one({"id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(404, "Trip not found")
    members = trip["members"]
    net = {m["id"]: 0.0 for m in members}
    weight_map = {m["id"]: _weight_of_member(m) for m in members}

    expenses = await db.expenses.find({"trip_id": trip_id, "kind": "expense"}, {"_id": 0}).to_list(5000)
    for e in expenses:
        split_ids = e["split_member_ids"] or [m["id"] for m in members]
        snap = e.get("weight_snapshots") or {}
        def wt(sid: str) -> int:
            if sid in snap:
                return int(snap[sid])
            return weight_map.get(sid, 1)
        total_weight = sum(wt(sid) for sid in split_ids)
        if total_weight == 0:
            continue
        per_unit = e["amount"] / total_weight
        for sid in split_ids:
            net[sid] = net.get(sid, 0) - per_unit * wt(sid)
        net[e["paid_by_member_id"]] = net.get(e["paid_by_member_id"], 0) + e["amount"]

    # apply settlements
    settlements = await db.settlements.find({"trip_id": trip_id}, {"_id": 0}).to_list(5000)
    for s in settlements:
        net[s["from_member_id"]] = net.get(s["from_member_id"], 0) + s["amount"]
        net[s["to_member_id"]] = net.get(s["to_member_id"], 0) - s["amount"]

    # round
    for k in net:
        net[k] = round(net[k], 2)

    # greedy settlement suggestion
    debtors = sorted([(mid, v) for mid, v in net.items() if v < -0.01], key=lambda x: x[1])
    creditors = sorted([(mid, v) for mid, v in net.items() if v > 0.01], key=lambda x: -x[1])
    transfers = []
    i = j = 0
    d = list(debtors); c = list(creditors)
    while i < len(d) and j < len(c):
        owe = -d[i][1]
        receive = c[j][1]
        pay = min(owe, receive)
        if pay > 0.01:
            transfers.append({"from_member_id": d[i][0], "to_member_id": c[j][0],
                              "amount": round(pay, 2)})
        d[i] = (d[i][0], d[i][1] + pay)
        c[j] = (c[j][0], c[j][1] - pay)
        if abs(d[i][1]) < 0.01:
            i += 1
        if abs(c[j][1]) < 0.01:
            j += 1
    return {"net": net, "transfers": transfers, "members": members,
            "currency": trip.get("currency", "INR"),
            "per_person": [
                {"member_id": m["id"], "member_name": m["name"], "kind": m["kind"],
                 "people_count": _weight_of_member(m),
                 "net_total": net.get(m["id"], 0.0),
                 "net_per_person": round(net.get(m["id"], 0.0) / _weight_of_member(m), 2),
                 "family_members": m.get("family_members", []) }
                for m in members
            ]}


@api.get("/trips/{trip_id}/balances")
async def balances(trip_id: str, user=Depends(get_current_user)):
    await _trip_or_404(trip_id, user["id"])
    return await _compute_balances(trip_id)


@api.post("/trips/{trip_id}/settle")
async def settle(trip_id: str, body: SettleIn, user=Depends(get_current_user)):
    await _trip_or_404(trip_id, user["id"])
    doc = {"id": gen_id(), "trip_id": trip_id,
           "from_member_id": body.from_member_id,
           "to_member_id": body.to_member_id,
           "amount": float(body.amount),
           "created_at": now_utc().isoformat(),
           "created_by": user["id"]}
    await db.settlements.insert_one(doc)
    doc.pop("_id", None)
    return doc


# ---------- Reports ----------
@api.get("/trips/{trip_id}/report")
async def report(trip_id: str, user=Depends(get_current_user)):
    trip = await _trip_or_404(trip_id, user["id"])
    expenses = await db.expenses.find({"trip_id": trip_id}, {"_id": 0}).to_list(5000)
    bal = await _compute_balances(trip_id)
    # category breakdown
    by_cat = {}
    by_date = {}
    total_expense = 0.0
    total_income = 0.0
    for e in expenses:
        if e["kind"] == "expense":
            by_cat[e["category"]] = by_cat.get(e["category"], 0) + e["amount"]
            by_date[e["date"]] = by_date.get(e["date"], 0) + e["amount"]
            total_expense += e["amount"]
        else:
            total_income += e["amount"]
    return {
        "trip": trip,
        "total_expense": round(total_expense, 2),
        "total_income": round(total_income, 2),
        "budget": trip.get("budget"),
        "by_category": [{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()],
        "by_date": [{"date": k, "amount": round(v, 2)} for k, v in sorted(by_date.items())],
        "balances": bal,
    }


@api.get("/trips/{trip_id}/report.xlsx")
async def report_xlsx(trip_id: str, token: str,
                      _unused=None):
    # token in query for easy mobile download
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(401, "User not found")
    trip = await _trip_or_404(trip_id, user["id"])
    expenses = await db.expenses.find({"trip_id": trip_id}, {"_id": 0}).to_list(5000)
    bal = await _compute_balances(trip_id)
    members_by_id = {m["id"]: m for m in trip["members"]}

    wb = Workbook()

    # Sheet 1 – Summary
    s1 = wb.active
    s1.title = "Summary"
    s1["A1"] = f"Trip: {trip['name']}"
    s1["A1"].font = Font(bold=True, size=16)
    s1["A2"] = f"Date: {trip['travel_date']}   Currency: {trip.get('currency','INR')}"
    s1["A3"] = f"Budget: {trip.get('budget', 'N/A')}"
    total = sum(e["amount"] for e in expenses if e["kind"] == "expense")
    s1["A4"] = f"Total Spent: {round(total,2)}"

    # Sheet 2 – By Category
    s2 = wb.create_sheet("By Category")
    s2.append(["Category", "Amount"])
    by_cat = {}
    for e in expenses:
        if e["kind"] == "expense":
            by_cat[e["category"]] = by_cat.get(e["category"], 0) + e["amount"]
    for k, v in by_cat.items():
        s2.append([k, round(v, 2)])

    # Sheet 3 – Per Member
    s3 = wb.create_sheet("Per Member")
    s3.append(["Member", "Type", "People", "Net Balance", "Per-Person"])
    for pp in bal["per_person"]:
        s3.append([pp["member_name"], pp["kind"], pp["people_count"],
                   pp["net_total"], pp["net_per_person"]])

    # Sheet 3b – Family per-person breakdown
    s3b = wb.create_sheet("Per Family Person")
    s3b.append(["Family", "Person", "Share of Net Balance"])
    for pp in bal["per_person"]:
        if pp["kind"] == "family" and pp["family_members"]:
            for name in pp["family_members"]:
                s3b.append([pp["member_name"], name, pp["net_per_person"]])

    # Sheet 4 – Transactions
    s4 = wb.create_sheet("Transactions")
    s4.append(["Date", "Kind", "Category", "Description", "Amount", "Paid By", "Split Among"])
    for e in sorted(expenses, key=lambda x: x.get("date", "")):
        paid = members_by_id.get(e["paid_by_member_id"], {}).get("name", "?")
        split = ", ".join(members_by_id.get(sid, {}).get("name", "?") for sid in e["split_member_ids"])
        s4.append([e["date"], e["kind"], e["category"], e.get("description", ""),
                   e["amount"], paid, split])

    # header styling
    header_fill = PatternFill("solid", fgColor="1C3F39")
    for sheet in [s2, s3, s3b, s4]:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{trip['name'].replace(' ','_')}_report.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ---------- Meta ----------
@api.get("/meta/categories")
async def get_categories():
    return CATEGORIES


# ---------- Startup ----------
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.trips.create_index("code", unique=True)
    await db.expenses.create_index([("trip_id", 1), ("created_at", -1)])
    # seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@trip.app")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    admin_pin = os.environ.get("ADMIN_PIN", "1234")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": gen_id(), "email": admin_email, "name": "Admin",
            "password_hash": hash_secret(admin_password),
            "pin_hash": hash_secret(admin_pin),
            "role": "admin", "created_at": now_utc().isoformat(),
        })
        logger.info("Seeded admin user")


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
