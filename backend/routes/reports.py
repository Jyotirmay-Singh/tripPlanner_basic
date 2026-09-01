import io

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import db
from utils.deps import get_current_user, _trip_or_404
from utils.date_rules import ensure_date_range, trip_date_label
from utils.balances import _compute_balances
from utils.display_names import member_display_names
from utils.ist_time import format_ist
from utils.security import decode_token
from services.report_builder import (
    build_expense_member_rows,
    build_members_families_rows,
    build_spend_reconciliation,
    build_split_math_rows,
    composition_label,
    entity_ledger_components,
    settle_adj_by_entity,
)
from services.report_pdf import build_report_pdf

router = APIRouter()

# ---------- XLSX styling (Phase 16) ----------
_BRAND = "1C3F39"
_HEADER_FILL = PatternFill("solid", fgColor=_BRAND)
_SUBSECTION_FILL = PatternFill("solid", fgColor="DDE9E5")
_REIMBURSEMENT_FILL = PatternFill("solid", fgColor="EEF3F1")
_TOTAL_FILL = PatternFill("solid", fgColor="EAF0EE")
_NET_FILL = PatternFill("solid", fgColor="D7E5E0")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color=_BRAND)
_BOLD = Font(bold=True)
_RIGHT = Alignment(horizontal="right")
_THIN_BRAND = Side(style="thin", color=_BRAND)
_MEDIUM_BRAND = Side(style="medium", color=_BRAND)
# Thousands separator, 2dp, negatives in red parentheses (professional accounting format).
_MONEY_FMT = "#,##0.00;[Red](#,##0.00)"


def _style_header_row(ws, row: int, ncols: int) -> None:
    """Bold white text on the brand fill across a header row."""
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL


def _money(cell) -> None:
    """Right-align a numeric cell and apply the currency number format."""
    cell.number_format = _MONEY_FMT
    cell.alignment = _RIGHT


def _set_widths(ws, widths) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_subsection_row(ws, row: int, ncols: int, *, reimbursement: bool = False) -> None:
    """Label a reconciliation subsection without relying only on colour."""
    fill = _REIMBURSEMENT_FILL if reimbursement else _SUBSECTION_FILL
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BOLD
        cell.fill = fill


def _style_summary_total(ws, row: int, ncols: int, *, strong: bool = False) -> None:
    """Style a subtotal or the final net row across its full table width."""
    side = _MEDIUM_BRAND if strong else _THIN_BRAND
    border = Border(top=side, bottom=side if strong else Side())
    fill = _NET_FILL if strong else _TOTAL_FILL
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BOLD
        cell.fill = fill
        cell.border = border


async def _load_report_expenses(trip_id: str) -> list:
    """Load the complete transaction set for reports; never silently truncate reimbursements."""
    return await db.expenses.find({"trip_id": trip_id}, {"_id": 0}).to_list(length=None)


# ---------- Reports ----------
@router.get("/trips/{trip_id}/report")
async def report(trip_id: str, user=Depends(get_current_user)):
    trip = ensure_date_range(await _trip_or_404(trip_id, user["id"]))
    expenses = await _load_report_expenses(trip_id)
    bal = await _compute_balances(trip_id)
    # category breakdown — signed amounts net together (a refund reduces its category + the total).
    by_cat = {}
    by_date = {}
    total_expense = 0.0
    for e in expenses:
        by_cat[e["category"]] = by_cat.get(e["category"], 0) + e["amount"]
        by_date[e["date"]] = by_date.get(e["date"], 0) + e["amount"]
        total_expense += e["amount"]
    return {
        "trip": trip,
        "total_expense": round(total_expense, 2),
        "budget": trip.get("budget"),
        "by_category": [{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()],
        "by_date": [{"date": k, "amount": round(v, 2)} for k, v in sorted(by_date.items())],
        "balances": bal,
    }


@router.get("/trips/{trip_id}/report.xlsx")
async def report_xlsx(trip_id: str, token: str,
                      _unused=None):
    # token in query for easy mobile download
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(401, "User not found")
    trip = await _trip_or_404(trip_id, user["id"])
    expenses = await _load_report_expenses(trip_id)
    bal = await _compute_balances(trip_id)
    # Disambiguated top-level labels (rule a + families) — one source of truth shared with the app.
    display = member_display_names(trip["members"])

    members = trip["members"]
    cur = trip.get("currency", "INR")
    reconciliation = build_spend_reconciliation(members, expenses)
    # The SAME two overlays _compute_balances applies to `net` (read-only; no engine change): the
    # non-pending settlements AND every Phase-20 payment. Both must feed the Settlements column or it
    # diverges from the ledger `net` the sheet reconciles against. Fetched once and reused (the
    # Payments tab below renders the same `payments` list).
    settlements = await db.settlements.find(
        {"trip_id": trip_id, "status": {"$ne": "pending"}}, {"_id": 0}).to_list(5000)
    payments = await db.payments.find({"trip_id": trip_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(5000)

    wb = Workbook()

    # ----- Tab 1: Summary (trip header + spend-by-entity + by-category) -----
    s1 = wb.active
    s1.title = "Summary"
    s1["A1"] = trip["name"]
    s1["A1"].font = Font(bold=True, size=16)
    row = 2
    for label, value in [("Dates", trip_date_label(trip)),
                         ("Share code", trip.get("code", "")),
                         ("Currency", cur),
                         ("Members", composition_label(members))]:
        s1.cell(row=row, column=1, value=label).font = _BOLD
        s1.cell(row=row, column=2, value=value)
        row += 1
    s1.cell(row=row, column=1, value="Budget").font = _BOLD
    if trip.get("budget") is not None:
        _money(s1.cell(row=row, column=2, value=round(trip["budget"], 2)))
    else:
        s1.cell(row=row, column=2, value="N/A")
    row += 1
    s1.cell(row=row, column=1, value="Net spend").font = _BOLD
    _money(s1.cell(row=row, column=2, value=reconciliation["totals"]["net"]))
    row += 2

    # One reconciliation model feeds both dimensions: gross, reimbursements, then net.
    s1.cell(row=row, column=1, value="Net spend by entity").font = _TITLE_FONT
    row += 1
    s1.cell(row=row, column=1, value="GROSS SPEND")
    _style_subsection_row(s1, row, 3)
    row += 1
    s1.cell(row=row, column=1, value="Entity")
    s1.cell(row=row, column=2, value="Type")
    s1.cell(row=row, column=3, value=f"Amount ({cur})")
    _style_header_row(s1, row, 3)
    row += 1
    gross_entities = [
        item for item in reconciliation["entities"]
        if item["gross"] != 0 or item["entity_id"] is not None
    ]
    for item in gross_entities:
        s1.cell(row=row, column=1, value=item["name"])
        s1.cell(row=row, column=2, value=item["type"])
        _money(s1.cell(row=row, column=3, value=item["gross"]))
        row += 1
    s1.cell(row=row, column=1, value="Gross spend subtotal")
    _money(s1.cell(row=row, column=3, value=reconciliation["totals"]["gross"]))
    _style_summary_total(s1, row, 3)
    row += 2

    s1.cell(row=row, column=1, value="REIMBURSEMENTS")
    _style_subsection_row(s1, row, 3, reimbursement=True)
    row += 1
    s1.cell(row=row, column=1, value="Entity")
    s1.cell(row=row, column=2, value="Type")
    s1.cell(row=row, column=3, value=f"Amount ({cur})")
    _style_header_row(s1, row, 3)
    row += 1
    reimbursement_entities = sorted(
        (item for item in reconciliation["entities"] if item["reimbursements"] != 0),
        key=lambda item: (-item["reimbursements"], item["name"]),
    )
    for item in reimbursement_entities:
        s1.cell(row=row, column=1, value=item["name"])
        s1.cell(row=row, column=2, value=item["type"])
        _money(s1.cell(row=row, column=3, value=item["reimbursements"]))
        row += 1
    s1.cell(row=row, column=1, value="Total reimbursements")
    _money(s1.cell(row=row, column=3, value=reconciliation["totals"]["reimbursements"]))
    _style_summary_total(s1, row, 3)
    row += 1
    s1.cell(row=row, column=1, value="Net spend")
    _money(s1.cell(row=row, column=3, value=reconciliation["totals"]["net"]))
    _style_summary_total(s1, row, 3, strong=True)
    row += 3

    s1.cell(row=row, column=1, value="Net spend by category").font = _TITLE_FONT
    row += 1
    s1.cell(row=row, column=1, value="GROSS SPEND")
    _style_subsection_row(s1, row, 2)
    row += 1
    s1.cell(row=row, column=1, value="Category")
    s1.cell(row=row, column=2, value=f"Amount ({cur})")
    _style_header_row(s1, row, 2)
    row += 1
    for item in reconciliation["categories"]:
        if item["gross"] == 0:
            continue
        s1.cell(row=row, column=1, value=item["category"])
        _money(s1.cell(row=row, column=2, value=item["gross"]))
        row += 1
    s1.cell(row=row, column=1, value="Gross spend subtotal")
    _money(s1.cell(row=row, column=2, value=reconciliation["totals"]["gross"]))
    _style_summary_total(s1, row, 2)
    row += 2

    s1.cell(row=row, column=1, value="REIMBURSEMENTS")
    _style_subsection_row(s1, row, 2, reimbursement=True)
    row += 1
    s1.cell(row=row, column=1, value="Category")
    s1.cell(row=row, column=2, value=f"Amount ({cur})")
    _style_header_row(s1, row, 2)
    row += 1
    for item in reconciliation["categories"]:
        if item["reimbursements"] == 0:
            continue
        s1.cell(row=row, column=1, value=item["category"])
        _money(s1.cell(row=row, column=2, value=item["reimbursements"]))
        row += 1
    s1.cell(row=row, column=1, value="Total reimbursements")
    _money(s1.cell(row=row, column=2, value=reconciliation["totals"]["reimbursements"]))
    _style_summary_total(s1, row, 2)
    row += 1
    s1.cell(row=row, column=1, value="Net spend")
    _money(s1.cell(row=row, column=2, value=reconciliation["totals"]["net"]))
    _style_summary_total(s1, row, 2, strong=True)
    _set_widths(s1, [28, 22, 18])

    # ----- Tab 2: Members & Families (Paid | Share | Settlements | Net, reconciling) -----
    s2 = wb.create_sheet("Members & Families")
    mf_headers = ["Name", "Type", "Family", f"Gross Spent ({cur})", f"Share of Expenses ({cur})",
                  f"Settlements ({cur})", f"Net Balance ({cur})"]
    s2.append(mf_headers)
    _style_header_row(s2, 1, len(mf_headers))
    paid_map, _ = entity_ledger_components(expenses, members)
    # settlements + payments = the exact overlay set _compute_balances lays over `net`, so the
    # Settlements column matches the ledger (partial payments included) and Net still foots.
    settle_map = settle_adj_by_entity(settlements + payments)
    for mf in build_members_families_rows(bal["per_person"], paid_map, settle_map, display):
        s2.append([mf["name"], mf["type"], mf["family"],
                   mf["paid"] if mf["paid"] is not None else "—",
                   mf["share"] if mf["share"] is not None else "—",
                   mf["settle"] if mf["settle"] is not None else "—",
                   mf["net"] if mf["net"] is not None else "—"])
        rr = s2.max_row
        for col in range(4, 8):
            c = s2.cell(row=rr, column=col)
            if isinstance(c.value, (int, float)):
                _money(c)
            else:
                c.alignment = _RIGHT
        if mf["kind"] in ("family_subtotal", "total"):
            for col in range(1, len(mf_headers) + 1):
                s2.cell(row=rr, column=col).font = _BOLD
        elif mf["kind"] == "family_member":
            s2.cell(row=rr, column=1).alignment = Alignment(indent=2)
    s2.freeze_panes = "A2"
    _set_widths(s2, [22, 14, 18, 16, 18, 16, 16])

    # ----- Tab 3: Split Math (flagship — one auditable block per expense) -----
    s3 = wb.create_sheet("Split Math")
    sm_headers = ["Expense", "Date", "Total Amount", "Split Mode", "Participant",
                  "Participant Type", "Units", f"Per-Unit Cost ({cur})", f"Allocated ({cur})"]
    s3.append(sm_headers)
    _style_header_row(s3, 1, len(sm_headers))
    for blk in build_split_math_rows(expenses, members):
        amt = round(blk["amount"], 2)
        for p in blk["participants"]:
            s3.append([blk["expense"], blk["date"], amt, blk["mode"], p["participant"],
                       p["ptype"], p["units"], round(p["per_unit"], 2), round(p["allocated"], 2)])
            rr = s3.max_row
            _money(s3.cell(row=rr, column=3))
            s3.cell(row=rr, column=7).alignment = _RIGHT
            _money(s3.cell(row=rr, column=8))
            _money(s3.cell(row=rr, column=9))
        s3.append([f"{blk['expense']} — Subtotal", "", amt, blk["mode"], "", "",
                   blk["subtotal_units"], "", blk["subtotal_allocated"]])
        rr = s3.max_row
        for col in range(1, len(sm_headers) + 1):
            s3.cell(row=rr, column=col).font = _BOLD
        _money(s3.cell(row=rr, column=3))
        s3.cell(row=rr, column=7).alignment = _RIGHT
        _money(s3.cell(row=rr, column=9))
    s3.freeze_panes = "A2"
    _set_widths(s3, [24, 18, 14, 12, 20, 16, 8, 16, 16])

    # ----- Tab 4: Transactions (exploded: one row per member + a per-person pivot) -----
    # Each expense expands into ONE ROW PER TRIP MEMBER showing that member's share ("Total Payable").
    # Amount / Split Mode / Paid By print once per expense (top row of the block); non-participants
    # show "-". A right-side pivot totals each person; the bottom row grand-totals Amount and Total
    # Payable. All figures come from build_expense_member_rows (reuses the ledger split math).
    s4 = wb.create_sheet("Transactions")
    tx = build_expense_member_rows(expenses, members)
    tx_headers = [
        "Sr No", "Category", "Description", "Date", f"Canonical Amount ({cur})",
        "Original Amount", "Original Currency", "Exchange Rate", "Effective Rate Date",
        "Rate Provider", "Rate Mode", "Original EXACT Allocations", "Split Mode", "Paid By",
        "Family", "Person Name", f"Total Payable ({cur})",
    ]
    s4.append(tx_headers)
    _style_header_row(s4, 1, len(tx_headers))
    for blk in tx["blocks"]:
        for i, r in enumerate(blk["rows"]):
            first = i == 0  # Amount / Split Mode / Paid By only on the block's first member row
            s4.append([
                blk["sr_no"] if first else None,
                blk["category"] if first else None,
                blk["description"] if first else None,
                blk["date"] if first else None,
                round(blk["amount"], 2) if first else None,
                round(blk["original_amount"], 2) if first else None,
                blk["original_currency"] if first else None,
                blk["exchange_rate"] if first else None,
                blk["exchange_rate_date"] if first else None,
                blk["exchange_rate_provider"] if first else None,
                blk["exchange_rate_mode"] if first else None,
                blk["original_exact_allocations"] if first else None,
                blk["mode"] if first else None,
                blk["paid_by"] if first else None,
                r["family"], r["person"],
                r["payable"] if r["participates"] else "-",
            ])
            rr = s4.max_row
            if first:
                _money(s4.cell(row=rr, column=5))
                _money(s4.cell(row=rr, column=6))
            pc = s4.cell(row=rr, column=17)
            if r["participates"]:
                _money(pc)
            else:
                pc.alignment = _RIGHT
    # Grand Total row (Sum(Amount) == Sum(Total Payable))
    s4.append(["Grand Total", None, None, None, tx["grand_amount"], None, None, None, None,
               None, None, None, None, None, None, None, tx["grand_payable"]])
    gr = s4.max_row
    for col in (1, 5, 17):
        s4.cell(row=gr, column=col).font = _BOLD
    _money(s4.cell(row=gr, column=5))
    _money(s4.cell(row=gr, column=17))

    # Right-side pivot (Person Name | Sum of Total Payable), one blank column after the main table.
    PV_NAME, PV_SUM = 19, 20
    s4.cell(row=1, column=PV_NAME, value="Person Name")
    s4.cell(row=1, column=PV_SUM, value=f"Sum of Total Payable ({cur})")
    for c in (PV_NAME, PV_SUM):
        hc = s4.cell(row=1, column=c)
        hc.font = _HEADER_FONT
        hc.fill = _HEADER_FILL
    pr = 2
    for prow in tx["pivot"]["rows"]:
        s4.cell(row=pr, column=PV_NAME, value=prow["name"])
        _money(s4.cell(row=pr, column=PV_SUM, value=prow["total"]))
        pr += 1
    s4.cell(row=pr, column=PV_NAME, value="Grand Total").font = _BOLD
    gt = s4.cell(row=pr, column=PV_SUM, value=tx["pivot"]["grand_total"])
    _money(gt)
    gt.font = _BOLD

    s4.freeze_panes = "A2"
    _set_widths(s4, [
        8, 14, 20, 16, 18, 16, 14, 16, 18, 24, 12, 42, 12, 16, 16, 16, 18, 4, 18, 20,
    ])

    # ----- Tab 5: Payments (Phase 20) — every recorded (partial) payment, one row each -----
    # A flat log of the settle-up payments: three partial payments = three rows. Names are the same
    # disambiguated labels as the rest of the report; Amount uses the trip currency; a bold Total row
    # sums the column. Display-only — these payments already offset the ledger via _compute_balances.
    s5 = wb.create_sheet("Payments")
    pay_headers = ["Payer", "Receiver", f"Amount ({cur})", "Date & Time", "Remark"]
    s5.append(pay_headers)
    _style_header_row(s5, 1, len(pay_headers))
    pay_total = 0.0
    for p in payments:
        dt_label = format_ist(p.get("created_at"))  # stored UTC -> IST display (Phase 24)
        s5.append([display.get(p["from_member_id"], "?"), display.get(p["to_member_id"], "?"),
                   round(p["amount"], 2), dt_label, (p.get("note") or "").strip() or "—"])
        _money(s5.cell(row=s5.max_row, column=3))
        pay_total += round(p["amount"], 2)
    s5.append(["Total", "", round(pay_total, 2), "", ""])
    tr = s5.max_row
    for col in range(1, len(pay_headers) + 1):
        s5.cell(row=tr, column=col).font = _BOLD
    _money(s5.cell(row=tr, column=3))
    s5.freeze_panes = "A2"
    _set_widths(s5, [24, 24, 16, 20, 30])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{trip['name'].replace(' ','_')}_report.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/trips/{trip_id}/report.pdf")
async def report_pdf(trip_id: str, token: str,
                     _unused=None):
    # Additive PDF variant of the report — same ?token= auth as report.xlsx (opened via a browser
    # link, so the JWT rides on the query string, not a header). Renders the FULL four-section report
    # (Summary, Members & Families, exploded Transactions, Payments) from the SAME pure builders the
    # spreadsheet uses, so the two can never diverge in value.
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(401, "User not found")
    trip = await _trip_or_404(trip_id, user["id"])
    members = trip["members"]
    expenses = await _load_report_expenses(trip_id)
    reconciliation = build_spend_reconciliation(members, expenses)
    payments = await db.payments.find({"trip_id": trip_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(5000)
    # Members & Families rows — identical construction to the XLSX route (same builders + the same
    # settlements + payments overlay), so the PDF's Settlements column and reconciliation match.
    bal = await _compute_balances(trip_id)
    display = member_display_names(members)
    settlements = await db.settlements.find(
        {"trip_id": trip_id, "status": {"$ne": "pending"}}, {"_id": 0}).to_list(5000)
    paid_map, _ = entity_ledger_components(expenses, members)
    settle_map = settle_adj_by_entity(settlements + payments)
    mf_rows = build_members_families_rows(bal["per_person"], paid_map, settle_map, display)
    pdf_bytes = build_report_pdf(
        trip, members, expenses, trip.get("currency", "INR"),
        reconciliation=reconciliation, payments=payments, mf_rows=mf_rows,
    )
    fname = f"{trip['name'].replace(' ','_')}_report.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
