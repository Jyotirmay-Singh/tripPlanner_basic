# Phase 20: partial payments (ledger offset, validation, RBAC, persistence, Excel reconciliation).
# Live API tests (mirror test_settlements.py — they hit a running server via requests). The pure
# roll-up + RBAC-predicate unit tests live in test_payments_rollup.py (no server needed).
import io
import os
import uuid

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'http://localhost:8000').rstrip('/')


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _register(api_client):
    """Register a fresh user; return {token, id, email}."""
    email = f"test_{uuid.uuid4().hex[:8]}@gmail.com"
    resp = api_client.post(f"{BASE_URL}/api/auth/register", json={
        "email": email, "password": "test12345", "pin": "4321", "name": "Pay Tester",
    })
    if resp.status_code != 200:
        pytest.skip(f"User registration failed: {resp.status_code}")
    data = resp.json()
    return {"token": data["access_token"], "id": data["user"]["id"], "email": email}


def _make_trip_two_members(api_client, owner_token):
    """Owner creates a trip; a second user (userB) joins as an individual.

    Returns (trip_id, owner_member_id, b_member_id, userB). Owner's member is the admin root; userB's
    is a plain non-admin member linked to userB's account (a valid receiver).
    """
    trip = api_client.post(f"{BASE_URL}/api/trips", json={
        "name": "TEST_Payments Trip",
        "start_date": "2026-01-10", "end_date": "2026-01-15", "currency": "INR",
    }, headers=_auth(owner_token)).json()
    trip_id = trip["id"]
    owner_member_id = trip["members"][0]["id"]
    user_b = _register(api_client)
    joined = api_client.post(f"{BASE_URL}/api/trips/join", json={
        "code": trip["code"], "mode": "individual",
    }, headers=_auth(user_b["token"])).json()
    b_member_id = next(m["id"] for m in joined["members"] if m.get("user_id") == user_b["id"])
    return trip_id, owner_member_id, b_member_id, user_b


def _add_expense(api_client, token, trip_id, paid_by, amount=200.0):
    api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
        "kind": "expense", "amount": amount, "category": "Food", "description": "Dinner",
        "date": "11-03-27", "paid_by_member_id": paid_by, "split_member_ids": [],
    }, headers=_auth(token))


def _balances(api_client, token, trip_id):
    return api_client.get(f"{BASE_URL}/api/trips/{trip_id}/balances", headers=_auth(token)).json()


def _transfer_amount(bal, from_id, to_id):
    for t in bal["transfers"]:
        if t["from_member_id"] == from_id and t["to_member_id"] == to_id:
            return t["amount"]
    return 0.0


class TestPaymentLedger:
    def test_partial_payment_reduces_pair_payable(self, api_client, test_user):
        # Owner paid 200 split between two -> m_b owes m_owner 100. Greedy m_b->m_owner 100.
        trip_id, m_owner, m_b, _b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_owner, amount=200.0)
        before = _balances(api_client, test_user["token"], trip_id)
        assert abs(_transfer_amount(before, m_b, m_owner) - 100.0) < 0.01

        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 60.0,
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 200, resp.text
        doc = resp.json()
        assert doc["amount"] == 60.0 and doc["currency"] == "INR"
        assert doc["from_member_id"] == m_b and doc["to_member_id"] == m_owner

        after = _balances(api_client, test_user["token"], trip_id)
        # Headline shrinks 100 -> 40; net moves toward 0.
        assert abs(_transfer_amount(after, m_b, m_owner) - 40.0) < 0.01
        assert abs(after["net"][m_owner] - 40.0) < 0.01
        assert abs(after["net"][m_b] - (-40.0)) < 0.01

    def test_payment_persists_and_offsets_after_new_expense(self, api_client, test_user):
        # Worked example: after a payment, a NEW expense flips the direction; the payment persists.
        trip_id, m_owner, m_b, _b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_owner, amount=200.0)
        # m_b owes m_owner 100; record 60 (owner is admin -> may record).
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 60.0,
        }, headers=_auth(test_user["token"]))
        # New expense: m_b pays 200 split both -> combined ledger nets to 0 before the payment; the
        # persisted 60 payment then makes m_owner owe m_b 60.
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_b, amount=200.0)
        after = _balances(api_client, test_user["token"], trip_id)
        assert abs(after["net"][m_owner] - (-60.0)) < 0.01
        assert abs(after["net"][m_b] - 60.0) < 0.01
        assert abs(_transfer_amount(after, m_owner, m_b) - 60.0) < 0.01


class TestPaymentValidation:
    def _setup_debt(self, api_client, token):
        # m_b owes m_owner 100 (greedy m_b -> m_owner).
        trip_id, m_owner, m_b, user_b = _make_trip_two_members(api_client, token)
        _add_expense(api_client, token, trip_id, paid_by=m_owner, amount=200.0)
        return trip_id, m_owner, m_b, user_b

    def test_rejects_non_positive_amount(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = self._setup_debt(api_client, test_user["token"])
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 0.0,
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 400

    def test_rejects_from_equals_to(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = self._setup_debt(api_client, test_user["token"])
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_owner, "to_member_id": m_owner, "amount": 10.0,
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 400

    def test_rejects_unknown_member(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = self._setup_debt(api_client, test_user["token"])
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": "ghost", "amount": 10.0,
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 400

    def test_rejects_non_suggested_pair(self, api_client, test_user):
        # Reverse direction (m_owner->m_b) is not suggested; must be rejected.
        trip_id, m_owner, m_b, _b = self._setup_debt(api_client, test_user["token"])
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_owner, "to_member_id": m_b, "amount": 10.0,
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 400

    def test_rejects_overpayment(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = self._setup_debt(api_client, test_user["token"])
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 150.0,  # payable is 100
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 400


class TestPaymentRBAC:
    def test_non_admin_receiver_can_record(self, api_client, test_user):
        # m_b (userB) paid -> m_owner owes m_b. userB is the RECEIVER and may record though non-admin.
        trip_id, m_owner, m_b, user_b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_b, amount=200.0)
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_owner, "to_member_id": m_b, "amount": 50.0,
        }, headers=_auth(user_b["token"]))
        assert resp.status_code == 200, resp.text

    def test_admin_can_record(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_b, amount=200.0)
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_owner, "to_member_id": m_b, "amount": 50.0,
        }, headers=_auth(test_user["token"]))
        assert resp.status_code == 200, resp.text

    def test_payer_cannot_self_record(self, api_client, test_user):
        # m_owner paid -> m_b owes m_owner. userB is the PAYER (from=m_b) and must be blocked.
        trip_id, m_owner, m_b, user_b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_owner, amount=200.0)
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 50.0,
        }, headers=_auth(user_b["token"]))
        assert resp.status_code == 403

    def test_unrelated_member_cannot_record(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_owner, amount=200.0)
        user_c = _register(api_client)
        trip = api_client.get(f"{BASE_URL}/api/trips/{trip_id}", headers=_auth(test_user["token"])).json()
        api_client.post(f"{BASE_URL}/api/trips/join", json={"code": trip["code"], "mode": "individual"},
                        headers=_auth(user_c["token"]))
        resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 50.0,
        }, headers=_auth(user_c["token"]))
        assert resp.status_code == 403

    def test_edit_and_delete_gated_to_receiver_or_admin(self, api_client, test_user):
        # m_b paid -> m_owner owes m_b; userB (receiver) records, then a bystander is blocked on
        # PATCH/DELETE while the receiver succeeds.
        trip_id, m_owner, m_b, user_b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_b, amount=200.0)
        pid = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_owner, "to_member_id": m_b, "amount": 40.0,
        }, headers=_auth(user_b["token"])).json()["id"]

        user_c = _register(api_client)
        trip = api_client.get(f"{BASE_URL}/api/trips/{trip_id}", headers=_auth(test_user["token"])).json()
        api_client.post(f"{BASE_URL}/api/trips/join", json={"code": trip["code"], "mode": "individual"},
                        headers=_auth(user_c["token"]))
        assert api_client.patch(f"{BASE_URL}/api/trips/{trip_id}/payments/{pid}", json={"amount": 30.0},
                                headers=_auth(user_c["token"])).status_code == 403
        assert api_client.delete(f"{BASE_URL}/api/trips/{trip_id}/payments/{pid}",
                                 headers=_auth(user_c["token"])).status_code == 403
        # Receiver edits then deletes.
        assert api_client.patch(f"{BASE_URL}/api/trips/{trip_id}/payments/{pid}", json={"amount": 30.0},
                                headers=_auth(user_b["token"])).status_code == 200
        assert api_client.delete(f"{BASE_URL}/api/trips/{trip_id}/payments/{pid}",
                                 headers=_auth(user_b["token"])).status_code == 200

    def test_non_member_blocked_everywhere(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_owner, amount=200.0)
        pid = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 20.0,
        }, headers=_auth(test_user["token"])).json()["id"]
        outsider = _register(api_client)
        h = _auth(outsider["token"])
        assert api_client.get(f"{BASE_URL}/api/trips/{trip_id}/payments", headers=h).status_code == 403
        assert api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m_b, "to_member_id": m_owner, "amount": 5.0,
        }, headers=h).status_code == 403
        assert api_client.patch(f"{BASE_URL}/api/trips/{trip_id}/payments/{pid}", json={"amount": 5.0},
                                headers=h).status_code == 403
        assert api_client.delete(f"{BASE_URL}/api/trips/{trip_id}/payments/{pid}",
                                 headers=h).status_code == 403


class TestPaymentExcelReconciles:
    def test_payments_tab_rows_reconcile(self, api_client, test_user):
        trip_id, m_owner, m_b, _b = _make_trip_two_members(api_client, test_user["token"])
        _add_expense(api_client, test_user["token"], trip_id, paid_by=m_owner, amount=200.0)
        amounts = [30.0, 25.0]
        for amt in amounts:
            r = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
                "from_member_id": m_b, "to_member_id": m_owner, "amount": amt,
            }, headers=_auth(test_user["token"]))
            assert r.status_code == 200, r.text

        token = test_user["token"]
        resp = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/report.xlsx?token={token}")
        assert resp.status_code == 200, resp.text
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Payments" in wb.sheetnames
        ws = wb["Payments"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows, "expected at least the payment rows + total"
        total_row = rows[-1]
        data_rows = rows[:-1]
        assert total_row[0] == "Total"
        assert len(data_rows) == len(amounts)
        assert abs(sum(r[2] for r in data_rows) - sum(amounts)) < 0.01
        assert abs(total_row[2] - sum(amounts)) < 0.01
