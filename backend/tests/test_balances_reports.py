# Balances, settle-up, and reports tests
import io
import pytest
import requests
import os
from openpyxl import load_workbook

BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'http://localhost:8000').rstrip('/')

class TestBalances:
    """Balance calculation and settle-up tests"""

    def test_get_balances(self, api_client, test_user):
        """Test GET /trips/{id}/balances returns net + transfers + members"""
        # Create trip with 2 members
        trip_resp = api_client.post(f"{BASE_URL}/api/trips", json={
            "name": "TEST_Balance Trip",
            "start_date": "2026-01-10", "end_date": "2026-01-15",
            "currency": "INR"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        trip_id = trip_resp.json()["id"]
        member1_id = trip_resp.json()["members"][0]["id"]

        member2_resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/members", json={
            "name": "TEST_Member2",
            "kind": "individual"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        member2_id = member2_resp.json()["id"]

        # Add expense
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense",
            "amount": 200.0,
            "category": "Food",
            "description": "Lunch",
            "date": "11-03-27",
            "paid_by_member_id": member1_id,
            "split_member_ids": []
        }, headers={"Authorization": f"Bearer {test_user['token']}"})

        # Get balances
        response = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/balances", headers={
            "Authorization": f"Bearer {test_user['token']}"
        })
        assert response.status_code == 200
        data = response.json()
        assert "net" in data
        assert "transfers" in data
        assert "members" in data
        assert "currency" in data
        
        # Member1 paid 200, owes 100, net = +100
        # Member2 paid 0, owes 100, net = -100
        assert abs(data["net"][member1_id] - 100.0) < 0.01
        assert abs(data["net"][member2_id] - (-100.0)) < 0.01
        
        # Should have 1 transfer suggestion
        assert len(data["transfers"]) == 1
        assert data["transfers"][0]["from_member_id"] == member2_id
        assert data["transfers"][0]["to_member_id"] == member1_id
        assert abs(data["transfers"][0]["amount"] - 100.0) < 0.01

    def test_settle_up(self, api_client, test_user):
        """Test POST /trips/{id}/settle decreases debtor's debt"""
        # Create trip with 2 members
        trip_resp = api_client.post(f"{BASE_URL}/api/trips", json={
            "name": "TEST_Settle Trip",
            "start_date": "2026-01-10", "end_date": "2026-01-15",
            "currency": "USD"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        trip_id = trip_resp.json()["id"]
        member1_id = trip_resp.json()["members"][0]["id"]

        member2_resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/members", json={
            "name": "TEST_Member2",
            "kind": "individual"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        member2_id = member2_resp.json()["id"]

        # Add expense
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense",
            "amount": 300.0,
            "category": "Travel",
            "description": "Taxi",
            "date": "16-04-27",
            "paid_by_member_id": member1_id,
            "split_member_ids": []
        }, headers={"Authorization": f"Bearer {test_user['token']}"})

        # Settle up: member2 pays member1 150
        response = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/settle", json={
            "from_member_id": member2_id,
            "to_member_id": member1_id,
            "amount": 150.0
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        assert response.status_code == 200
        data = response.json()
        assert data["amount"] == 150.0
        assert data["from_member_id"] == member2_id
        assert data["to_member_id"] == member1_id

        # Check balances after settlement
        balance_resp = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/balances", headers={
            "Authorization": f"Bearer {test_user['token']}"
        })
        balances = balance_resp.json()
        
        # After settlement: member1 net = 150 - 150 = 0, member2 net = -150 + 150 = 0
        # Wait, member1 paid 300, owes 150, net = +150
        # member2 paid 0, owes 150, net = -150
        # After settle: member2 pays 150 to member1
        # member1 net = 150 + 150 = 300, member2 net = -150 + 150 = 0
        # Actually: net[member1] = paid - owed = 300 - 150 = 150
        # After settlement: member2 pays 150, so member2 net = -150 + 150 = 0
        # member1 receives 150, so member1 net = 150 - 150 = 0
        assert abs(balances["net"][member1_id]) < 0.01
        assert abs(balances["net"][member2_id]) < 0.01


class TestReports:
    """Report generation tests"""

    def test_get_report_json(self, api_client, test_user):
        """Test GET /trips/{id}/report returns JSON summary"""
        # Create trip with expenses
        trip_resp = api_client.post(f"{BASE_URL}/api/trips", json={
            "name": "TEST_Report Trip",
            "start_date": "2026-01-10", "end_date": "2026-01-15",
            "budget": 2000.0,
            "currency": "EUR"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        trip_id = trip_resp.json()["id"]
        member_id = trip_resp.json()["members"][0]["id"]

        # Add expenses in different categories
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense",
            "amount": 500.0,
            "category": "Food",
            "description": "Meals",
            "date": "21-05-27",
            "paid_by_member_id": member_id,
            "split_member_ids": []
        }, headers={"Authorization": f"Bearer {test_user['token']}"})

        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense",
            "amount": 800.0,
            "category": "Accommodation",
            "description": "Hotel",
            "date": "22-05-27",
            "paid_by_member_id": member_id,
            "split_member_ids": []
        }, headers={"Authorization": f"Bearer {test_user['token']}"})

        # Get report
        response = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/report", headers={
            "Authorization": f"Bearer {test_user['token']}"
        })
        assert response.status_code == 200
        data = response.json()
        assert "trip" in data
        assert "total_expense" in data
        assert "total_income" not in data  # income concept removed (signed-amount model)
        assert "budget" in data
        assert "by_category" in data
        assert "by_date" in data
        assert "balances" in data
        
        assert data["total_expense"] == 1300.0
        assert data["budget"] == 2000.0
        assert len(data["by_category"]) == 2
        assert any(c["category"] == "Food" and c["amount"] == 500.0 for c in data["by_category"])
        assert any(c["category"] == "Accommodation" and c["amount"] == 800.0 for c in data["by_category"])

    def test_get_report_xlsx(self, api_client, test_user):
        """Test GET /trips/{id}/report.xlsx?token=... returns XLSX binary"""
        # Create trip with expenses
        trip_resp = api_client.post(f"{BASE_URL}/api/trips", json={
            "name": "TEST_XLSX Report Trip",
            "start_date": "2026-01-10", "end_date": "2026-01-15",
            "currency": "GBP"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        trip_id = trip_resp.json()["id"]
        member_id = trip_resp.json()["members"][0]["id"]

        # A family member so the PER_FAMILY math tab has a multi-entity line item.
        fam_resp = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/members", json={
            "name": "TEST_Fam", "kind": "family", "family_members": ["a", "b"]
        }, headers={"Authorization": f"Bearer {test_user['token']}"})
        fam_id = fam_resp.json()["id"]

        # PER_CAPITA line item -> feeds the Per-Capita Math tab.
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense",
            "amount": 150.0,
            "category": "Shopping",
            "description": "Souvenirs",
            "date": "26-06-27",
            "paid_by_member_id": member_id,
            "split_member_ids": [],
            "split_mode": "PER_CAPITA"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})

        # PER_FAMILY line item -> feeds the Per-Family Math tab.
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense",
            "amount": 120.0,
            "category": "Travel",
            "description": "Taxi",
            "date": "27-06-27",
            "paid_by_member_id": member_id,
            "split_member_ids": [member_id, fam_id],
            "split_mode": "PER_FAMILY"
        }, headers={"Authorization": f"Bearer {test_user['token']}"})

        # Get XLSX report
        response = api_client.get(
            f"{BASE_URL}/api/trips/{trip_id}/report.xlsx?token={test_user['token']}"
        )
        assert response.status_code == 200
        assert "spreadsheet" in response.headers.get("Content-Type", "")
        assert len(response.content) > 0
        # Verify it's a valid XLSX file (starts with PK zip signature)
        assert response.content[:2] == b'PK'

        # Phase 16: 4 professional tabs; Phase 20 appends a 5th "Payments" tab (in order).
        wb = load_workbook(io.BytesIO(response.content))
        assert wb.sheetnames == ["Summary", "Members & Families", "Split Math", "Transactions", "Payments"]

        # Summary carries the trip header, the per-entity Gross Spent block, and By category.
        summary_vals = [str(c.value) for rowcells in wb["Summary"].iter_rows() for c in rowcells]
        assert any("Gross Spent" in v for v in summary_vals)
        assert any(v == "Total Spent" for v in summary_vals)
        assert any(v == "By category" for v in summary_vals)

        # Members & Families header has the four reconciling money columns.
        mf = wb["Members & Families"]
        mf_header = [c.value for c in mf[1]]
        assert mf_header[0] == "Name" and mf_header[1] == "Type" and mf_header[2] == "Family"
        assert "Gross Spent" in mf_header[3] and "Share of Expenses" in mf_header[4]
        assert "Settlements" in mf_header[5] and "Net Balance" in mf_header[6]
        # Every numeric entity/total row reconciles: Net = Paid - Share + Settlements (to the cent).
        recon_rows = 0
        for name, typ, fam, paid, share, settle, net in mf.iter_rows(min_row=2, values_only=True):
            if all(isinstance(x, (int, float)) for x in (paid, share, settle, net)):
                assert abs(round(paid - share + settle, 2) - net) <= 0.011
                recon_rows += 1
        assert recon_rows >= 2  # at least the individual entity + the TOTAL row

        # Split Math is the combined per-(expense x participant) tab with subtotals.
        sm = wb["Split Math"]
        sm_header = [c.value for c in sm[1]]
        assert sm_header[:6] == ["Expense", "Date", "Total Amount", "Split Mode",
                                 "Participant", "Participant Type"]
        assert sm_header[6] == "Units" and "Per-Unit Cost" in sm_header[7] and "Allocated" in sm_header[8]
        # At least one expense subtotal whose Allocated equals the expense Total Amount.
        subtotal_checks = 0
        for expense, date, total_amt, mode, participant, ptype, units, per_unit, allocated in \
                sm.iter_rows(min_row=2, values_only=True):
            if isinstance(expense, str) and expense.endswith("Subtotal"):
                assert abs(allocated - total_amt) <= 0.011
                subtotal_checks += 1
        assert subtotal_checks >= 2  # one per expense (PER_CAPITA + PER_FAMILY line items)

        # Transactions journal (Phase 18 exploded layout): Split Mode is column 6, humanized.
        tx_header = [c.value for c in wb["Transactions"][1]]
        assert tx_header[5] == "Split Mode"
        tx_modes = {row[5] for row in wb["Transactions"].iter_rows(min_row=2, values_only=True) if row[5]}
        assert tx_modes <= {"Per-Person", "Per-Family"}

    def test_report_settlements_column_includes_partial_payments(self, api_client, test_user):
        """Phase 23 (BUG): the Members & Families 'Settlements' column must reflect BOTH the legacy
        non-pending settlements AND Phase-20 partial payments (the SAME overlay the ledger applies),
        or it understates reality. Pre-fix this cell read the settlement only. Also verifies the
        Payments header renamed to 'Receiver' and the full PDF renders."""
        h = {"Authorization": f"Bearer {test_user['token']}"}
        token = test_user["token"]

        trip = api_client.post(f"{BASE_URL}/api/trips", json={
            "name": "TEST_Phase23 Settle+Pay", "start_date": "2026-01-10",
            "end_date": "2026-01-15", "currency": "INR"}, headers=h).json()
        trip_id = trip["id"]
        m1 = trip["members"][0]["id"]  # creator's member (payer / receiver of the settlement)
        m2 = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/members",
                             json={"name": "TEST_Debtor", "kind": "individual"},
                             headers=h).json()["id"]

        # m1 fronts 200 split between the two -> m2 owes m1 100 (suggested transfer m2 -> m1).
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense", "amount": 200.0, "category": "Food", "description": "Dinner",
            "date": "11-03-27", "paid_by_member_id": m1, "split_member_ids": []}, headers=h)

        # Legacy settlement 30 (stamped paid) THEN a partial payment 15, both m2 -> m1.
        r_settle = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/settle", json={
            "from_member_id": m2, "to_member_id": m1, "amount": 30.0}, headers=h)
        assert r_settle.status_code == 200
        r_pay = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": m2, "to_member_id": m1, "amount": 15.0}, headers=h)
        assert r_pay.status_code == 200, r_pay.text

        # ----- XLSX: Members & Families Settlements column must include BOTH (30 + 15 = 45) -----
        resp = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/report.xlsx?token={token}")
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        mf = wb["Members & Families"]
        entity_rows, total_row = [], None
        for name, typ, fam, paid, share, settle, net in mf.iter_rows(min_row=2, values_only=True):
            if name == "TOTAL":
                total_row = (paid, share, settle, net)
            elif all(isinstance(x, (int, float)) for x in (paid, share, settle, net)):
                entity_rows.append({"paid": paid, "share": share, "settle": settle, "net": net})

        payer = next(r for r in entity_rows if abs(r["paid"] - 200.0) < 0.01)
        debtor = next(r for r in entity_rows if abs(r["paid"] - 0.0) < 0.01)
        # Debtor paid OUT 30 + 15 = 45 (would be ONLY 30 before the fix); receiver mirror -45.
        assert abs(debtor["settle"] - 45.0) < 0.01, debtor
        assert abs(payer["settle"] - (-45.0)) < 0.01, payer
        # Share is the TRUE engine allocation (each owes 100), NOT contaminated by the payment.
        assert abs(debtor["share"] - 100.0) < 0.01 and abs(payer["share"] - 100.0) < 0.01
        # Net foots with payments included: payer +55, debtor -55.
        assert abs(payer["net"] - 55.0) < 0.01 and abs(debtor["net"] - (-55.0)) < 0.01
        for r in entity_rows:
            assert abs(round(r["paid"] - r["share"] + r["settle"], 2) - r["net"]) <= 0.011
        # Totals: Σ Settlements == Σ Net == 0.
        assert total_row is not None
        assert abs(total_row[2]) < 0.011 and abs(total_row[3]) < 0.011

        # Payments tab header renamed Payee -> Receiver (Phase 23); Remark column added (settle-up note).
        assert [c.value for c in wb["Payments"][1]] == ["Payer", "Receiver", "Amount (INR)",
                                                        "Date & Time", "Remark"]

        # ----- PDF: full report renders (same builders -> same values) -----
        pdf = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/report.pdf?token={token}")
        assert pdf.status_code == 200
        assert "application/pdf" in pdf.headers.get("Content-Type", "")
        assert pdf.content[:4] == b"%PDF"
        assert len(pdf.content) > 0
