"""Route-level regression tests for report data loading."""

import asyncio
import io
from types import SimpleNamespace
from unittest.mock import AsyncMock

from openpyxl import load_workbook

from routes import reports


class _Cursor:
    def __init__(self, rows):
        self.rows = rows
        self.requested_length = "not-called"

    async def to_list(self, length=None):
        self.requested_length = length
        return self.rows

    def sort(self, *args):
        return self


class _ExpensesCollection:
    def __init__(self, cursor):
        self.cursor = cursor
        self.find_args = None

    def find(self, query, projection):
        self.find_args = (query, projection)
        return self.cursor


def test_load_report_expenses_uses_unbounded_cursor(monkeypatch):
    rows = [{"id": "gross"}, {"id": "reimbursement"}]
    cursor = _Cursor(rows)
    expenses = _ExpensesCollection(cursor)
    monkeypatch.setattr(reports, "db", SimpleNamespace(expenses=expenses))

    assert asyncio.run(reports._load_report_expenses("trip-1")) == rows
    assert expenses.find_args == ({"trip_id": "trip-1"}, {"_id": 0})
    assert cursor.requested_length is None


def test_xlsx_summary_renders_time_gross_reimbursement_and_net_separately(monkeypatch):
    members = [{
        "id": "time",
        "name": "Time",
        "kind": "family",
        "family_members": ["Hour", "Minute"],
        "family_member_ids": ["time:0", "time:1"],
    }]
    trip = {
        "id": "trip-1",
        "name": "Time report",
        "start_date": "2026-01-01",
        "end_date": "2026-01-02",
        "currency": "INR",
        "code": "TIME",
        "members": members,
    }
    expenses = [
        {
            "id": "gross",
            "amount": 14_000.0,
            "category": "Local Transportation",
            "description": "Taxi",
            "date": "01-01-26",
            "paid_by_member_id": "time",
            "split_member_ids": ["time"],
            "split_mode": "PER_FAMILY",
        },
        {
            "id": "reimbursement",
            "amount": -4_000.0,
            "category": "Local Transportation",
            "description": "Taxi refund",
            "date": "02-01-26",
            "paid_by_member_id": "time",
            "split_member_ids": ["time"],
            "split_mode": "PER_FAMILY",
        },
    ]
    per_person = [{
        "member_id": "time",
        "member_name": "Time",
        "kind": "family",
        "net_total": 0.0,
        "members": [
            {"id": "time:0", "name": "Hour", "net": 0.0},
            {"id": "time:1", "name": "Minute", "net": 0.0},
        ],
    }]

    monkeypatch.setattr(reports, "decode_token", lambda _token: {"sub": "user-1"})
    monkeypatch.setattr(reports, "_trip_or_404", AsyncMock(return_value=trip))
    monkeypatch.setattr(reports, "_load_report_expenses", AsyncMock(return_value=expenses))
    monkeypatch.setattr(
        reports, "_compute_balances", AsyncMock(return_value={"per_person": per_person}),
    )
    monkeypatch.setattr(reports, "db", SimpleNamespace(
        users=SimpleNamespace(find_one=AsyncMock(return_value={"id": "user-1"})),
        settlements=SimpleNamespace(find=lambda *_args, **_kwargs: _Cursor([])),
        payments=SimpleNamespace(find=lambda *_args, **_kwargs: _Cursor([])),
    ))

    async def render():
        response = await reports.report_xlsx("trip-1", "token")
        chunks = [chunk async for chunk in response.body_iterator]
        return response, b"".join(chunks)

    response, payload = asyncio.run(render())
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    summary = workbook["Summary"]
    rows = [tuple(cell.value for cell in row[:3]) for row in summary.iter_rows()]

    assert [row[0] for row in rows].count("GROSS SPEND") == 2
    assert [row[0] for row in rows].count("REIMBURSEMENTS") == 2
    assert [row[0] for row in rows].count("Gross spend subtotal") == 2
    assert [row[0] for row in rows].count("Total reimbursements") == 2
    assert [row[2] for row in rows if row[0] == "Time" and isinstance(row[2], (int, float))] == [
        14_000.0, 4_000.0,
    ]
    assert [row[1] for row in rows
            if row[0] == "Local Transportation" and isinstance(row[1], (int, float))] == [
        14_000.0, 4_000.0,
    ]
    assert [next(value for value in row[1:] if isinstance(value, (int, float)))
            for row in rows if row[0] == "Gross spend subtotal"] == [14_000.0, 14_000.0]
    assert [next(value for value in row[1:] if isinstance(value, (int, float)))
            for row in rows if row[0] == "Total reimbursements"] == [4_000.0, 4_000.0]
    assert [next(value for value in row[1:] if isinstance(value, (int, float)))
            for row in rows if row[0] == "Net spend"] == [10_000.0, 10_000.0, 10_000.0]
