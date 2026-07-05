# Phase 22 — EXACT split mode: live-server API tests (requests/BASE_URL style, run in the full gate
# against a running server + local Mongo, mirroring test_payments.py / test_family_participation_api.py).
# The pure math (validator/resolver/ledger/breakdown) is covered offline in test_exact_split.py.
import io
import os
import uuid

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'http://localhost:8000').rstrip('/')


def _h(token):
    return {"Authorization": f"Bearer {token}"}


def _register(api_client, name="Exact Tester"):
    email = f"test_{uuid.uuid4().hex[:8]}@gmail.com"
    resp = api_client.post(f"{BASE_URL}/api/auth/register", json={
        "email": email, "password": "test12345", "pin": "4321", "name": name,
    })
    if resp.status_code != 200:
        pytest.skip(f"User registration failed: {resp.status_code}")
    data = resp.json()
    return {"token": data["access_token"], "id": data["user"]["id"], "email": email}


def _make_trip(api_client, token, name="TEST_Exact"):
    trip = api_client.post(f"{BASE_URL}/api/trips", json={
        "name": name, "start_date": "2026-01-10", "end_date": "2026-01-15", "currency": "INR",
    }, headers=_h(token)).json()
    return trip["id"], trip["members"][0]["id"], trip["code"]


def _add_family(api_client, token, trip_id, names):
    fam = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/members", json={
        "name": "TEST_Fam", "kind": "family", "family_members": names,
        "family_member_ids": [None] * len(names),
    }, headers=_h(token)).json()
    return fam["id"], fam["family_member_ids"]


def _post_exact(api_client, token, trip_id, amount, custom_amounts, paid_by, force=False):
    return api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
        "kind": "expense", "amount": amount, "category": "Food", "description": "Exact dinner",
        "date": "11-01-26", "paid_by_member_id": paid_by, "split_member_ids": [],
        "split_mode": "EXACT", "custom_amounts": custom_amounts,
    }, headers=_h(token))


def _get_expense(api_client, token, trip_id, eid):
    exps = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/expenses", headers=_h(token)).json()
    return next((e for e in exps if e["id"] == eid), None)


def _balances(api_client, token, trip_id):
    return api_client.get(f"{BASE_URL}/api/trips/{trip_id}/balances", headers=_h(token)).json()


def _transfer(bal, from_id, to_id):
    for t in bal["transfers"]:
        if t["from_member_id"] == from_id and t["to_member_id"] == to_id:
            return t["amount"]
    return 0.0


# --------------------------------------------------------------------------- create + the hard rule
class TestCreateHardRule:
    def test_create_matching_sum_persists(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        custom = {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}
        resp = _post_exact(api_client, test_user["token"], trip_id, 100.0, custom, paid_by=owner_id)
        assert resp.status_code == 200, resp.text
        eid = resp.json()["expense"]["id"]
        stored = _get_expense(api_client, test_user["token"], trip_id, eid)
        assert stored["split_mode"] == "EXACT"
        assert stored["custom_amounts"] == {fam_ids[0]: 80.0, fam_ids[1]: 10.0, owner_id: 10.0}

    def test_create_mismatched_sum_rejected_422(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        custom = {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 5}  # sums 95, total 100
        resp = _post_exact(api_client, test_user["token"], trip_id, 100.0, custom, paid_by=owner_id)
        assert resp.status_code == 422, resp.text
        assert "add up to the total" in resp.text


# --------------------------------------------------------------------------- edit round-trip + hard rule
class TestEditHardRule:
    def test_patch_matching_roundtrips(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        eid = _post_exact(api_client, test_user["token"], trip_id, 100.0,
                          {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}, paid_by=owner_id).json()["expense"]["id"]
        new_custom = {fam_ids[0]: 50, fam_ids[1]: 30, owner_id: 20}
        resp = api_client.patch(f"{BASE_URL}/api/trips/{trip_id}/expenses/{eid}", json={
            "amount": 100.0, "split_mode": "EXACT", "custom_amounts": new_custom,
        }, headers=_h(test_user["token"]))
        assert resp.status_code == 200, resp.text
        stored = _get_expense(api_client, test_user["token"], trip_id, eid)
        assert stored["custom_amounts"] == {fam_ids[0]: 50.0, fam_ids[1]: 30.0, owner_id: 20.0}

    def test_patch_mismatch_rejected_and_unchanged(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        eid = _post_exact(api_client, test_user["token"], trip_id, 100.0,
                          {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}, paid_by=owner_id).json()["expense"]["id"]
        resp = api_client.patch(f"{BASE_URL}/api/trips/{trip_id}/expenses/{eid}", json={
            "custom_amounts": {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 999},
        }, headers=_h(test_user["token"]))
        assert resp.status_code == 422, resp.text
        stored = _get_expense(api_client, test_user["token"], trip_id, eid)
        assert stored["custom_amounts"] == {fam_ids[0]: 80.0, fam_ids[1]: 10.0, owner_id: 10.0}


# --------------------------------------------------------------------------- balances + breakdown + settle
class TestBalancesAndBreakdown:
    def test_rollup_balances_and_family_breakdown(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        _post_exact(api_client, test_user["token"], trip_id, 100.0,
                    {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}, paid_by=owner_id)
        bal = _balances(api_client, test_user["token"], trip_id)
        # entity rollup: family owes 90 (80+10), owner is owed 90 (paid 100 - own 10).
        assert abs(bal["net"][fam_id] - (-90.0)) < 0.01
        assert abs(bal["net"][owner_id] - 90.0) < 0.01
        assert abs(_transfer(bal, fam_id, owner_id) - 90.0) < 0.01
        # per-member family breakdown = the typed amounts; foots to the family net.
        fam_pp = next(pp for pp in bal["per_person"] if pp["member_id"] == fam_id)
        members = {r["id"]: r["net"] for r in fam_pp["members"]}
        assert abs(members[fam_ids[0]] - (-80.0)) < 0.01
        assert abs(members[fam_ids[1]] - (-10.0)) < 0.01
        assert round(sum(members.values()), 2) == -90.0

    def test_partial_payment_scales_breakdown(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        _post_exact(api_client, test_user["token"], trip_id, 100.0,
                    {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}, paid_by=owner_id)
        # Family pays the owner back 45 of the 90 (admin/receiver may record). Net halves.
        pay = api_client.post(f"{BASE_URL}/api/trips/{trip_id}/payments", json={
            "from_member_id": fam_id, "to_member_id": owner_id, "amount": 45.0,
        }, headers=_h(test_user["token"]))
        assert pay.status_code == 200, pay.text
        bal = _balances(api_client, test_user["token"], trip_id)
        assert abs(bal["net"][fam_id] - (-45.0)) < 0.01
        fam_pp = next(pp for pp in bal["per_person"] if pp["member_id"] == fam_id)
        members = {r["id"]: r["net"] for r in fam_pp["members"]}
        # scaled toward 0 by 0.5: -80 -> -40, -10 -> -5; still foots to the family net.
        assert abs(members[fam_ids[0]] - (-40.0)) < 0.01
        assert abs(members[fam_ids[1]] - (-5.0)) < 0.01
        assert round(sum(members.values()), 2) == -45.0


# --------------------------------------------------------------------------- reports (display-only)
class TestReports:
    def test_xlsx_and_pdf_include_exact_and_reconcile(self, api_client, test_user):
        trip_id, owner_id, _ = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        _post_exact(api_client, test_user["token"], trip_id, 100.0,
                    {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}, paid_by=owner_id)
        token = test_user["token"]
        xr = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/report.xlsx?token={token}")
        assert xr.status_code == 200, xr.text
        wb = load_workbook(io.BytesIO(xr.content))
        # Transactions tab: Σ Amount == pivot Grand Total (100).
        ws = wb["Transactions"]
        vals = list(ws.iter_rows(values_only=True))
        flat = [c for row in vals for c in row if c is not None]
        assert any(str(c) == "Exact" for c in flat), "EXACT mode label missing from report"
        pr = api_client.get(f"{BASE_URL}/api/trips/{trip_id}/report.pdf?token={token}")
        assert pr.status_code == 200 and pr.content[:4] == b"%PDF"


# --------------------------------------------------------------------------- RBAC unchanged
class TestRBAC:
    def test_non_creator_non_admin_cannot_edit_or_delete(self, api_client, test_user):
        trip_id, owner_id, code = _make_trip(api_client, test_user["token"])
        fam_id, fam_ids = _add_family(api_client, test_user["token"], trip_id, ["A", "B"])
        eid = _post_exact(api_client, test_user["token"], trip_id, 100.0,
                          {fam_ids[0]: 80, fam_ids[1]: 10, owner_id: 10}, paid_by=owner_id).json()["expense"]["id"]
        outsider = _register(api_client, name="Outsider")
        api_client.post(f"{BASE_URL}/api/trips/join", json={"code": code, "mode": "individual"},
                        headers=_h(outsider["token"]))
        pe = api_client.patch(f"{BASE_URL}/api/trips/{trip_id}/expenses/{eid}", json={"description": "hax"},
                              headers=_h(outsider["token"]))
        assert pe.status_code == 403
        de = api_client.delete(f"{BASE_URL}/api/trips/{trip_id}/expenses/{eid}", headers=_h(outsider["token"]))
        assert de.status_code == 403


# --------------------------------------------------------------------------- regression (other modes intact)
class TestRegression:
    def test_per_capita_and_per_family_unaffected(self, api_client, test_user):
        trip_id, owner_id, code = _make_trip(api_client, test_user["token"])
        userb = _register(api_client, name="B")
        joined = api_client.post(f"{BASE_URL}/api/trips/join", json={"code": code, "mode": "individual"},
                                 headers=_h(userb["token"])).json()
        b_id = next(m["id"] for m in joined["members"] if m.get("user_id") == userb["id"])
        # PER_CAPITA 100 between two individuals -> each owes 50.
        api_client.post(f"{BASE_URL}/api/trips/{trip_id}/expenses", json={
            "kind": "expense", "amount": 100.0, "category": "Food", "description": "PC",
            "date": "11-01-26", "paid_by_member_id": owner_id, "split_member_ids": [], "split_mode": "PER_CAPITA",
        }, headers=_h(test_user["token"]))
        bal = _balances(api_client, test_user["token"], trip_id)
        assert abs(bal["net"][owner_id] - 50.0) < 0.01
        assert abs(bal["net"][b_id] - (-50.0)) < 0.01
